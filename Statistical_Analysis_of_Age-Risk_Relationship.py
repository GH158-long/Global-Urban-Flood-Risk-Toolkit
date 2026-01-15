#!/usr/bin/env python3
# -*- coding: utf-8 -*-


import sys
import os
import gc
import time
import threading
import traceback
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional
from collections import defaultdict

import numpy as np
import pandas as pd
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

# ============================================================================
# Configuration Constants
# ============================================================================

APP_NAME = "Urban Flood Comprehensive Analyzer"
APP_VERSION = "3.1.0"
DEFAULT_INPUT = r"E:\Global_Flood\FIG\DATA\Urban_age_V17.parquet"
DEFAULT_OUTPUT = r"E:\Global_Flood\FIG\DATA"

# Analysis Dimensions
AGE_CATEGORIES = [5, 10, 15, 20, 25, 30, 35, 40, 45, 50]
POP_YEARS = [1975, 1980, 1985, 1990, 1995, 2000, 2005, 2010, 2015, 2020, 2025]

# Classification Mappings
DEVELOPED_MAP = {1: 'Developed', 2: 'Developing'}
INCOME_MAP = {
    0: 'Unclassified', 1: 'Low Income', 2: 'Lower Middle Income',
    3: 'Upper Middle Income', 4: 'High Income'
}
CONTINENT_MAP = {
    10: 'Asia', 11: 'Europe', 12: 'Africa', 13: 'North America',
    14: 'South America', 15: 'Oceania', 17: 'Seven Seas'
}
COUNTRY_MAP = {
    100: 'Indonesia', 101: 'Malaysia', 102: 'Chile', 103: 'Bolivia', 104: 'Peru',
    105: 'Argentina', 106: 'Zekeliya Barracks', 107: 'Cyprus', 108: 'India', 109: 'China',
    110: 'Israel', 111: 'Palestine', 112: 'Lebanon', 113: 'Ethiopia', 114: 'South Sudan',
    115: 'Somalia', 116: 'Kenya', 117: 'Malawi', 118: 'Tanzania', 119: 'Syria',
    120: 'Somaliland', 121: 'France', 122: 'Suriname', 123: 'Guyana', 124: 'South Korea',
    125: 'North Korea', 126: 'Morocco', 127: 'Western Sahara', 128: 'Costa Rica',
    129: 'Nicaragua', 130: 'Congo', 131: 'DR Congo', 132: 'Bhutan', 133: 'Ukraine',
    134: 'Belarus', 135: 'Namibia', 136: 'South Africa', 137: 'Saint Martin',
    138: 'Sint Maarten', 139: 'Oman', 140: 'Uzbekistan', 141: 'Kazakhstan',
    142: 'Tajikistan', 143: 'Lithuania', 144: 'Brazil', 145: 'Uruguay', 146: 'Mongolia',
    147: 'Russia', 148: 'Czechia', 149: 'Germany', 150: 'Estonia', 151: 'Latvia',
    152: 'Norway', 153: 'Sweden', 154: 'Finland', 155: 'Vietnam', 156: 'Cambodia',
    157: 'Luxembourg', 158: 'UAE', 159: 'Belgium', 160: 'Georgia', 161: 'North Macedonia',
    162: 'Albania', 163: 'Azerbaijan', 164: 'Kosovo', 165: 'Turkey', 166: 'Spain',
    167: 'Laos', 168: 'Kyrgyzstan', 169: 'Armenia', 170: 'Denmark', 171: 'Libya',
    172: 'Tunisia', 173: 'Romania', 174: 'Hungary', 175: 'Slovakia', 176: 'Poland',
    177: 'Ireland', 178: 'United Kingdom', 179: 'Greece', 180: 'Zambia',
    181: 'Sierra Leone', 182: 'Guinea', 183: 'Liberia', 184: 'CAR', 185: 'Sudan',
    186: 'Djibouti', 187: 'Eritrea', 188: 'Austria', 189: 'Iraq', 190: 'Italy',
    191: 'Switzerland', 192: 'Iran', 193: 'Netherlands', 194: 'Liechtenstein',
    195: "Cote d'Ivoire", 196: 'Serbia', 197: 'Mali', 198: 'Senegal', 199: 'Nigeria',
    200: 'Benin', 201: 'Angola', 202: 'Croatia', 203: 'Slovenia', 204: 'Qatar',
    205: 'Saudi Arabia', 206: 'Botswana', 207: 'Zimbabwe', 208: 'Pakistan',
    209: 'Bulgaria', 210: 'Thailand', 211: 'San Marino', 212: 'Haiti',
    213: 'Dominican Rep.', 214: 'Chad', 215: 'Kuwait', 216: 'El Salvador',
    217: 'Guatemala', 218: 'Timor-Leste', 219: 'Brunei', 220: 'Monaco', 221: 'Algeria',
    222: 'Mozambique', 223: 'Eswatini', 224: 'Burundi', 225: 'Rwanda', 226: 'Myanmar',
    227: 'Bangladesh', 228: 'Andorra', 229: 'Afghanistan', 230: 'Montenegro',
    231: 'Bosnia & Herzegovina', 232: 'Uganda', 233: 'Guantanamo Bay', 234: 'Cuba',
    235: 'Honduras', 236: 'Ecuador', 237: 'Colombia', 238: 'Paraguay',
    239: 'Brazil Island', 240: 'Portugal', 241: 'Moldova', 242: 'Turkmenistan',
    243: 'Jordan', 244: 'Nepal', 245: 'Lesotho', 246: 'Cameroon', 247: 'Gabon',
    248: 'Niger', 249: 'Burkina Faso', 250: 'Togo', 251: 'Ghana', 252: 'Guinea-Bissau',
    253: 'Gibraltar', 254: 'USA', 255: 'Canada', 256: 'Mexico', 257: 'Belize',
    258: 'Panama', 259: 'Venezuela', 260: 'Papua New Guinea', 261: 'Egypt',
    262: 'Yemen', 263: 'Mauritania', 264: 'Equatorial Guinea', 265: 'Gambia',
    266: 'Hong Kong', 267: 'Vatican City', 268: 'N. Cyprus', 269: 'UN Buffer Zone',
    270: 'Siachen Glacier', 271: 'Baikonur', 272: 'Akrotiri SBA',
    273: 'S. Patagonia Ice Field', 274: 'Bir Tawil', 276: 'Australia',
    277: 'Greenland', 278: 'Fiji', 279: 'New Zealand', 280: 'New Caledonia',
    281: 'Madagascar', 282: 'Philippines', 283: 'Sri Lanka', 284: 'Curacao',
    285: 'Aruba', 286: 'Bahamas', 287: 'Turks & Caicos', 288: 'Taiwan',
    289: 'Japan', 290: 'St Pierre & Miquelon', 291: 'Iceland', 292: 'Pitcairn Islands',
    293: 'French Polynesia', 294: 'TAAF France', 295: 'Seychelles', 296: 'Kiribati',
    297: 'Marshall Islands', 298: 'Trinidad & Tobago', 299: 'Grenada',
    300: 'St Vincent', 301: 'Barbados', 302: 'Saint Lucia', 303: 'Dominica',
    304: 'US Minor Outlying Is.', 305: 'Montserrat', 306: 'Antigua & Barbuda',
    307: 'St Kitts & Nevis', 308: 'US Virgin Islands', 309: 'St Barthelemy',
    310: 'Puerto Rico', 311: 'Anguilla', 312: 'British Virgin Is.', 313: 'Jamaica',
    314: 'Cayman Islands', 315: 'Bermuda', 316: 'Heard & McDonald Is.',
    317: 'Saint Helena', 318: 'Mauritius', 319: 'Comoros', 320: 'Sao Tome & Principe',
    321: 'Cape Verde', 322: 'Malta', 323: 'Jersey', 324: 'Guernsey', 325: 'Isle of Man',
    326: 'Aland Finland', 327: 'Faroe Islands', 328: 'AU Indian Ocean Terr.',
    329: 'BIOT', 330: 'Singapore', 331: 'Norfolk Island', 332: 'Cook Islands',
    333: 'Tonga', 334: 'Wallis & Futuna', 335: 'Samoa', 336: 'Solomon Islands',
    337: 'Tuvalu', 338: 'Maldives', 339: 'Nauru', 340: 'Micronesia',
    342: 'Falkland Islands', 343: 'Vanuatu', 344: 'Niue NZ', 345: 'American Samoa',
    346: 'Palau', 347: 'Guam', 348: 'N. Mariana Islands', 349: 'Bahrain',
    350: 'Coral Sea Islands', 351: 'Wake Island', 352: 'Clipperton Island',
    353: 'Macau', 354: 'Ashmore & Cartier Is.', 355: 'Bajo Nuevo Bank',
    356: 'Serranilla Bank', 357: 'Scarborough Shoal'
}

CELL_AREA_KM2 = 0.01
ENV_COLUMNS = ['Hazard_Index_EAD', 'Slope', 'DEM', 'Distance_from_River']


# ============================================================================
# Theme Configuration
# ============================================================================

class Theme:
    PRIMARY = '#1a73e8'
    PRIMARY_DARK = '#1557b0'
    PRIMARY_LIGHT = '#e8f0fe'

    BG_MAIN = '#f8f9fa'
    BG_CARD = '#ffffff'
    BG_DARK = '#202124'
    BG_LOG = '#1e1e1e'

    TEXT_PRIMARY = '#202124'
    TEXT_SECONDARY = '#5f6368'
    TEXT_MUTED = '#80868b'
    TEXT_LIGHT = '#ffffff'

    SUCCESS = '#34a853'
    WARNING = '#fbbc04'
    ERROR = '#ea4335'
    INFO = '#4285f4'

    BORDER = '#dadce0'
    BORDER_LIGHT = '#e8eaed'

    PROGRESS_BG = '#e8eaed'
    PROGRESS_FG = '#1a73e8'


class Icons:
    FOLDER = "📁"
    FILE = "📄"
    SAVE = "💾"
    PLAY = "▶️"
    STOP = "⏹️"
    SUCCESS = "✅"
    ERROR = "❌"
    WARNING = "⚠️"
    INFO = "ℹ️"
    ROCKET = "🚀"
    CLOCK = "⏱️"
    CALC = "🔢"
    CHART = "📊"
    GLOBE = "🌍"
    MONEY = "💰"
    PEOPLE = "👥"
    BUILDING = "🏢"
    GEAR = "⚙️"
    CHECK = "✓"
    ARROW = "→"
    BULLET = "•"
    LAYER = "📑"
    DATA = "📋"
    STATS = "📈"
    MAP = "🗺️"


# ============================================================================
# Core Data Analyzer
# ============================================================================

class UnifiedDataAnalyzer:
    """Unified Data Analyzer - Integrates all statistical computations

    v3.1 Changes:
    - Added "All" category for each grouping type
    - Statistics now include n (count) and SE (standard error = σ/√n)
    """

    def __init__(self, log_callback=None, progress_callback=None, detail_callback=None):
        self.log = log_callback or print
        self.progress = progress_callback or (lambda x, y: None)
        self.detail = detail_callback or print
        self.should_stop = False

        # [A] Area Statistics
        self.area_by_dev_age = {}

        # [B] Risk Index Statistics - Added 'all' for each grouping
        self.risk_stats = {
            'by_dev': {},  # Will include 'All' + Developed/Developing
            'by_income': {},  # Will include 'All' + income levels
            'by_continent': {},  # Will include 'All' + continents
            'by_country': {}  # Will include 'All' + countries
        }

        # [C] Environmental Factor Statistics - Added 'all' for each grouping
        self.env_stats = {
            'by_dev': {},
            'by_income': {},
            'by_continent': {},
            'by_country': {}
        }

        # [D] Population Exposure Statistics
        self.exposure_totals = {
            'by_country_year': {},
            'by_income_year': {}
        }
        self.pop_exposure = {
            'global': {},
            'by_dev': {},
            'by_income': {},
            'by_continent': {},
            'by_country': {}
        }

        # Counters
        self.stats_counters = {
            'total_rows': 0,
            'valid_rows': 0,
            'countries_found': set(),
            'continents_found': set(),
            'income_levels_found': set(),
            'ages_found': set()
        }

    def reset(self):
        """Reset all statistics"""
        self.area_by_dev_age = {}
        for key in self.risk_stats:
            self.risk_stats[key] = {}
        for key in self.env_stats:
            self.env_stats[key] = {}
        for key in self.exposure_totals:
            self.exposure_totals[key] = {}
        for key in self.pop_exposure:
            self.pop_exposure[key] = {}
        self.stats_counters = {
            'total_rows': 0,
            'valid_rows': 0,
            'countries_found': set(),
            'continents_found': set(),
            'income_levels_found': set(),
            'ages_found': set()
        }
        self.should_stop = False

    def _init_stats(self):
        return {'sum': 0.0, 'sq_sum': 0.0, 'count': 0,
                'min': float('inf'), 'max': float('-inf')}

    def _init_pop_exp(self):
        return {'pop_sum': 0.0, 'exp_sum': 0.0}

    def _update_stats(self, stats: dict, values: np.ndarray):
        if len(values) == 0:
            return
        stats['sum'] += np.nansum(values)
        stats['sq_sum'] += np.nansum(values ** 2)
        stats['count'] += np.count_nonzero(~np.isnan(values))
        valid_vals = values[~np.isnan(values)]
        if len(valid_vals) > 0:
            stats['min'] = min(stats['min'], np.min(valid_vals))
            stats['max'] = max(stats['max'], np.max(valid_vals))

    def _finalize_stats(self, stats: dict) -> dict:
        """Finalize statistics with n, σ, and σ/√n (standard error)"""
        if stats['count'] == 0:
            return {'mean': np.nan, 'std': np.nan, 'min': np.nan,
                    'max': np.nan, 'count': 0, 'se': np.nan}
        n = stats['count']
        mean = stats['sum'] / n
        variance = max(0, (stats['sq_sum'] / n) - (mean ** 2))
        std = np.sqrt(variance)
        # Standard Error = σ / √n
        se = std / np.sqrt(n) if n > 0 else np.nan
        return {
            'mean': mean,
            'std': std,
            'min': stats['min'] if stats['min'] != float('inf') else np.nan,
            'max': stats['max'] if stats['max'] != float('-inf') else np.nan,
            'count': n,
            'se': se  # σ/√n - Standard Error
        }

    def process_chunk(self, df: pd.DataFrame, exclude_zero: bool = True) -> int:
        """Process data chunk"""
        if 'Built_Up_Age' not in df.columns:
            self.detail(f"    {Icons.WARNING} Missing Built_Up_Age column, skipping chunk")
            return 0

        df_valid = df[df['Built_Up_Age'].isin(AGE_CATEGORIES)].copy()
        if len(df_valid) == 0:
            self.detail(f"    {Icons.INFO} No valid Built-up Age data")
            return 0

        self.stats_counters['total_rows'] += len(df)
        self.stats_counters['valid_rows'] += len(df_valid)

        self.stats_counters['ages_found'].update(df_valid['Built_Up_Age'].unique())
        if 'Country' in df.columns:
            self.stats_counters['countries_found'].update(df_valid['Country'].dropna().unique())
        if 'Sovereig' in df.columns:
            self.stats_counters['continents_found'].update(df_valid['Sovereig'].dropna().unique())
        if 'Income_Classification' in df.columns:
            self.stats_counters['income_levels_found'].update(
                df_valid['Income_Classification'].dropna().unique())

        for age in AGE_CATEGORIES:
            age_mask = df_valid['Built_Up_Age'] == age
            age_data = df_valid[age_mask]
            if len(age_data) == 0:
                continue

            self._process_area_stats(age_data, age)
            self._process_risk_stats(age_data, age, exclude_zero)
            self._process_env_stats(age_data, age, exclude_zero)
            self._process_pop_exposure(age_data, age)

        self._process_yearly_exposure(df_valid)

        return len(df_valid)

    def _process_area_stats(self, data: pd.DataFrame, age: int):
        key_all = ('All', age)
        self.area_by_dev_age[key_all] = self.area_by_dev_age.get(key_all, 0) + len(data)

        if 'Developed' in data.columns:
            for dev_code, dev_name in DEVELOPED_MAP.items():
                count = (data['Developed'] == dev_code).sum()
                if count > 0:
                    key = (dev_name, age)
                    self.area_by_dev_age[key] = self.area_by_dev_age.get(key, 0) + count

    def _process_risk_stats(self, data: pd.DataFrame, age: int, exclude_zero: bool):
        """Process Risk Index statistics with 'All' category for each grouping"""
        if 'Risk_Index' not in data.columns:
            return

        def get_values(subset):
            vals = subset['Risk_Index'].values.astype(float)
            if exclude_zero:
                vals = np.where(vals == 0, np.nan, vals)
            return vals

        # === Process by Development Status (includes "All") ===
        if 'Developed' in data.columns:
            # First, add "All" category for this grouping
            key_all = ('All', age)
            if key_all not in self.risk_stats['by_dev']:
                self.risk_stats['by_dev'][key_all] = self._init_stats()
            self._update_stats(self.risk_stats['by_dev'][key_all], get_values(data))

            # Then individual categories
            for dev_code, dev_name in DEVELOPED_MAP.items():
                subset = data[data['Developed'] == dev_code]
                if len(subset) == 0:
                    continue
                key = (dev_name, age)
                if key not in self.risk_stats['by_dev']:
                    self.risk_stats['by_dev'][key] = self._init_stats()
                self._update_stats(self.risk_stats['by_dev'][key], get_values(subset))

        # === Process by Income Level (includes "All") ===
        if 'Income_Classification' in data.columns:
            # First, add "All" category
            key_all = ('All', age)
            if key_all not in self.risk_stats['by_income']:
                self.risk_stats['by_income'][key_all] = self._init_stats()
            self._update_stats(self.risk_stats['by_income'][key_all], get_values(data))

            # Then individual categories
            for income_code, income_name in INCOME_MAP.items():
                subset = data[data['Income_Classification'] == income_code]
                if len(subset) == 0:
                    continue
                key = (income_name, age)
                if key not in self.risk_stats['by_income']:
                    self.risk_stats['by_income'][key] = self._init_stats()
                self._update_stats(self.risk_stats['by_income'][key], get_values(subset))

        # === Process by Continent (includes "All") ===
        if 'Sovereig' in data.columns:
            # First, add "All" category
            key_all = ('All', age)
            if key_all not in self.risk_stats['by_continent']:
                self.risk_stats['by_continent'][key_all] = self._init_stats()
            self._update_stats(self.risk_stats['by_continent'][key_all], get_values(data))

            # Then individual categories
            for cont_code, cont_name in CONTINENT_MAP.items():
                subset = data[data['Sovereig'] == cont_code]
                if len(subset) == 0:
                    continue
                key = (cont_name, age)
                if key not in self.risk_stats['by_continent']:
                    self.risk_stats['by_continent'][key] = self._init_stats()
                self._update_stats(self.risk_stats['by_continent'][key], get_values(subset))

        # === Process by Country (includes "All") ===
        if 'Country' in data.columns:
            # First, add "All" category
            key_all = ('All', age)
            if key_all not in self.risk_stats['by_country']:
                self.risk_stats['by_country'][key_all] = self._init_stats()
            self._update_stats(self.risk_stats['by_country'][key_all], get_values(data))

            # Then individual countries
            for country_code in data['Country'].unique():
                if pd.isna(country_code):
                    continue
                country_name = COUNTRY_MAP.get(int(country_code), f'Country_{int(country_code)}')
                subset = data[data['Country'] == country_code]
                if len(subset) == 0:
                    continue
                key = (country_name, age)
                if key not in self.risk_stats['by_country']:
                    self.risk_stats['by_country'][key] = self._init_stats()
                self._update_stats(self.risk_stats['by_country'][key], get_values(subset))

    def _process_env_stats(self, data: pd.DataFrame, age: int, exclude_zero: bool):
        """Process Environmental Factor statistics with 'All' category for each grouping"""
        available_cols = [c for c in ENV_COLUMNS if c in data.columns]
        if not available_cols:
            return

        def get_values(subset, col):
            vals = subset[col].values.astype(float)
            if exclude_zero:
                vals = np.where(vals == 0, np.nan, vals)
            return vals

        def process_group(data_subset, key, target_dict):
            if key not in target_dict:
                target_dict[key] = {col: self._init_stats() for col in available_cols}
            for col in available_cols:
                self._update_stats(target_dict[key][col], get_values(data_subset, col))

        # === Process by Development Status (includes "All") ===
        if 'Developed' in data.columns:
            # First, add "All" category
            process_group(data, ('All', age), self.env_stats['by_dev'])

            # Then individual categories
            for dev_code, dev_name in DEVELOPED_MAP.items():
                subset = data[data['Developed'] == dev_code]
                if len(subset) > 0:
                    process_group(subset, (dev_name, age), self.env_stats['by_dev'])

        # === Process by Income Level (includes "All") ===
        if 'Income_Classification' in data.columns:
            # First, add "All" category
            process_group(data, ('All', age), self.env_stats['by_income'])

            # Then individual categories
            for income_code, income_name in INCOME_MAP.items():
                subset = data[data['Income_Classification'] == income_code]
                if len(subset) > 0:
                    process_group(subset, (income_name, age), self.env_stats['by_income'])

        # === Process by Continent (includes "All") ===
        if 'Sovereig' in data.columns:
            # First, add "All" category
            process_group(data, ('All', age), self.env_stats['by_continent'])

            # Then individual categories
            for cont_code, cont_name in CONTINENT_MAP.items():
                subset = data[data['Sovereig'] == cont_code]
                if len(subset) > 0:
                    process_group(subset, (cont_name, age), self.env_stats['by_continent'])

        # === Process by Country (includes "All") ===
        if 'Country' in data.columns:
            # First, add "All" category
            process_group(data, ('All', age), self.env_stats['by_country'])

            # Then individual countries
            for country_code in data['Country'].unique():
                if pd.isna(country_code):
                    continue
                country_name = COUNTRY_MAP.get(int(country_code), f'Country_{int(country_code)}')
                subset = data[data['Country'] == country_code]
                if len(subset) > 0:
                    process_group(subset, (country_name, age), self.env_stats['by_country'])

    def _process_pop_exposure(self, data: pd.DataFrame, age: int):
        if 'POP_2025' not in data.columns or 'POP_Exposure_2025' not in data.columns:
            return

        pop_col = 'POP_2025'
        exp_col = 'POP_Exposure_2025'

        def update_pop_exp(key, subset, target_dict):
            if key not in target_dict:
                target_dict[key] = self._init_pop_exp()
            target_dict[key]['pop_sum'] += subset[pop_col].fillna(0).sum()
            target_dict[key]['exp_sum'] += subset[exp_col].fillna(0).sum()

        update_pop_exp(age, data, self.pop_exposure['global'])

        if 'Developed' in data.columns:
            # Add "All" category
            update_pop_exp(('All', age), data, self.pop_exposure['by_dev'])
            for dev_code, dev_name in DEVELOPED_MAP.items():
                subset = data[data['Developed'] == dev_code]
                if len(subset) > 0:
                    update_pop_exp((dev_name, age), subset, self.pop_exposure['by_dev'])

        if 'Income_Classification' in data.columns:
            # Add "All" category
            update_pop_exp(('All', age), data, self.pop_exposure['by_income'])
            for income_code, income_name in INCOME_MAP.items():
                subset = data[data['Income_Classification'] == income_code]
                if len(subset) > 0:
                    update_pop_exp((income_name, age), subset, self.pop_exposure['by_income'])

        if 'Sovereig' in data.columns:
            # Add "All" category
            update_pop_exp(('All', age), data, self.pop_exposure['by_continent'])
            for cont_code, cont_name in CONTINENT_MAP.items():
                subset = data[data['Sovereig'] == cont_code]
                if len(subset) > 0:
                    update_pop_exp((cont_name, age), subset, self.pop_exposure['by_continent'])

        if 'Country' in data.columns:
            # Add "All" category
            update_pop_exp(('All', age), data, self.pop_exposure['by_country'])
            for country_code in data['Country'].unique():
                if pd.isna(country_code):
                    continue
                country_name = COUNTRY_MAP.get(int(country_code), f'Country_{int(country_code)}')
                subset = data[data['Country'] == country_code]
                if len(subset) > 0:
                    update_pop_exp((country_name, age), subset, self.pop_exposure['by_country'])

    def _process_yearly_exposure(self, data: pd.DataFrame):
        if 'Country' in data.columns:
            for year in POP_YEARS:
                col = f'POP_Exposure_{year}'
                if col not in data.columns:
                    continue
                for country_code in data['Country'].unique():
                    if pd.isna(country_code):
                        continue
                    country_name = COUNTRY_MAP.get(int(country_code), f'Country_{int(country_code)}')
                    total = data.loc[data['Country'] == country_code, col].sum()
                    key = (country_name, year)
                    self.exposure_totals['by_country_year'][key] = \
                        self.exposure_totals['by_country_year'].get(key, 0) + total

        if 'Income_Classification' in data.columns:
            for year in POP_YEARS:
                col = f'POP_Exposure_{year}'
                if col not in data.columns:
                    continue
                for income_code, income_name in INCOME_MAP.items():
                    total = data.loc[data['Income_Classification'] == income_code, col].sum()
                    key = (income_name, year)
                    self.exposure_totals['by_income_year'][key] = \
                        self.exposure_totals['by_income_year'].get(key, 0) + total

    def calculate_final_results(self) -> dict:
        """Calculate final results with n and σ/√n (SE)"""
        results = {
            'summary': {},
            'area': {},
            'risk': {},
            'environment': {},
            'exposure': {}
        }

        results['summary'] = {
            'total_rows': self.stats_counters['total_rows'],
            'valid_rows': self.stats_counters['valid_rows'],
            'countries_count': len(self.stats_counters['countries_found']),
            'continents_count': len(self.stats_counters['continents_found']),
            'income_levels_count': len(self.stats_counters['income_levels_found']),
            'ages_count': len(self.stats_counters['ages_found'])
        }

        # [A] Area Statistics
        area_list = []
        total_all = sum(v for k, v in self.area_by_dev_age.items() if k[0] == 'All')
        for key, count in self.area_by_dev_age.items():
            dev_name, age = key
            area_km2 = count * CELL_AREA_KM2
            if dev_name == 'All':
                ratio = (count / total_all * 100) if total_all > 0 else 0
            else:
                dev_total = sum(v for k, v in self.area_by_dev_age.items() if k[0] == dev_name)
                ratio = (count / dev_total * 100) if dev_total > 0 else 0
            area_list.append({
                'Category': dev_name, 'Built_Up_Age': age,
                'Grid_Count': count, 'Area_km2': area_km2, 'Ratio_Percent': ratio
            })
        results['area']['by_dev_age'] = sorted(area_list, key=lambda x: (
            0 if x['Category'] == 'All' else 1, x['Category'], x['Built_Up_Age']))

        # [B] Risk Index - Now includes n (count) and SE (σ/√n)
        for group_name, group_data in self.risk_stats.items():
            risk_list = []
            for key, stats in group_data.items():
                cat_name, age = key
                final = self._finalize_stats(stats)
                risk_list.append({
                    'Category': cat_name, 'Built_Up_Age': age,
                    'Risk_Mean': final['mean'], 'Risk_Std': final['std'],
                    'Risk_Min': final['min'], 'Risk_Max': final['max'],
                    'Sample_Count': final['count'],  # n
                    'Risk_SE': final['se']  # σ/√n (Standard Error)
                })
            # Sort with "All" first
            results['risk'][group_name] = sorted(risk_list, key=lambda x: (
                0 if x['Category'] == 'All' else 1, x['Category'], x['Built_Up_Age']))

        # [C] Environmental Factors - Now includes n and SE for each factor
        for group_name, group_data in self.env_stats.items():
            env_list = []
            for key, col_stats in group_data.items():
                cat_name, age = key
                row = {'Category': cat_name, 'Built_Up_Age': age}
                for col in ENV_COLUMNS:
                    if col in col_stats:
                        final = self._finalize_stats(col_stats[col])
                        row[f'{col}_Mean'] = final['mean']
                        row[f'{col}_Std'] = final['std']
                        row[f'{col}_n'] = final['count']  # n
                        row[f'{col}_SE'] = final['se']  # σ/√n
                    else:
                        row[f'{col}_Mean'] = np.nan
                        row[f'{col}_Std'] = np.nan
                        row[f'{col}_n'] = 0
                        row[f'{col}_SE'] = np.nan
                env_list.append(row)
            # Sort with "All" first
            results['environment'][group_name] = sorted(env_list, key=lambda x: (
                0 if x['Category'] == 'All' else 1, x['Category'], x['Built_Up_Age']))

        # [D] Population Exposure
        global_ratio = []
        for age in AGE_CATEGORIES:
            if age in self.pop_exposure['global']:
                d = self.pop_exposure['global'][age]
                ratio = (d['exp_sum'] / d['pop_sum'] * 100) if d['pop_sum'] > 0 else np.nan
                global_ratio.append({
                    'Built_Up_Age': age, 'Total_Population': d['pop_sum'],
                    'Total_Exposure': d['exp_sum'], 'Exposure_Ratio_Percent': ratio
                })
        results['exposure']['global_ratio'] = global_ratio

        for group_name in ['by_dev', 'by_income', 'by_continent', 'by_country']:
            ratio_list = []
            for key, d in self.pop_exposure[group_name].items():
                cat_name, age = key
                ratio = (d['exp_sum'] / d['pop_sum'] * 100) if d['pop_sum'] > 0 else np.nan
                ratio_list.append({
                    'Category': cat_name, 'Built_Up_Age': age,
                    'Total_Population': d['pop_sum'], 'Total_Exposure': d['exp_sum'],
                    'Exposure_Ratio_Percent': ratio
                })
            # Sort with "All" first
            results['exposure'][f'{group_name}_ratio'] = sorted(
                ratio_list, key=lambda x: (
                    0 if x['Category'] == 'All' else 1, x['Category'], x['Built_Up_Age']))

        country_yearly = defaultdict(dict)
        for (country, year), total in self.exposure_totals['by_country_year'].items():
            country_yearly[country][year] = total
        country_list = []
        for country, years in country_yearly.items():
            country_list.append({
                'Country': country,
                'Years': years,
                'Total_2025': years.get(2025, 0)
            })
        country_list.sort(key=lambda x: x['Total_2025'], reverse=True)
        results['exposure']['by_country_year'] = country_list

        income_yearly = defaultdict(dict)
        for (income, year), total in self.exposure_totals['by_income_year'].items():
            income_yearly[income][year] = total
        results['exposure']['by_income_year'] = dict(income_yearly)

        return results


# ============================================================================
# Excel Report Generator
# ============================================================================

class ExcelReportGenerator:
    """Excel Report Generator with professional formatting

    v3.1 Changes:
    - Added n (sample count) and SE (standard error) columns
    - "All" category included in each sheet
    """

    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self._setup_styles()

    def _setup_styles(self):
        self.colors = {
            'primary': '1a73e8',
            'primary_dark': '1557b0',
            'header': '2e5090',
            'header_text': 'ffffff',
            'alt_row': 'f8f9fa',
            'border': 'dadce0',
            'success': '34a853',
            'warning': 'fbbc04',
            'link': '1a73e8',
            'all_highlight': 'e8f0fe'  # Highlight for "All" rows
        }

        self.fonts = {
            'title': Font(bold=True, size=24, color='1a73e8'),
            'subtitle': Font(bold=True, size=16, color='5f6368'),
            'section': Font(bold=True, size=14, color='202124'),
            'header': Font(bold=True, size=11, color='ffffff'),
            'body': Font(size=10, color='202124'),
            'small': Font(size=9, color='5f6368'),
            'link': Font(size=10, color='1a73e8', underline='single'),
            'number': Font(size=10, name='Consolas'),
            'all_bold': Font(bold=True, size=10, color='1557b0')  # For "All" category
        }

        self.fills = {
            'header': PatternFill('solid', fgColor='2e5090'),
            'alt_row': PatternFill('solid', fgColor='f8f9fa'),
            'highlight': PatternFill('solid', fgColor='e8f0fe'),
            'success': PatternFill('solid', fgColor='e6f4ea'),
            'warning': PatternFill('solid', fgColor='fef7e0'),
            'all_row': PatternFill('solid', fgColor='e8f0fe')  # For "All" rows
        }

        thin_border = Side(style='thin', color='dadce0')
        self.border = Border(left=thin_border, right=thin_border,
                             top=thin_border, bottom=thin_border)

        self.align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.align_right = Alignment(horizontal='right', vertical='center')

    def generate_report(self, results: dict) -> str:
        wb = Workbook()

        self._create_overview_sheet(wb, results)
        self._create_area_sheet(wb, results['area'])
        self._create_risk_sheets(wb, results['risk'])
        self._create_env_sheets(wb, results['environment'])
        self._create_exposure_sheets(wb, results['exposure'])

        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb['Sheet']

        output_path = os.path.join(self.output_dir, f'Urban_Flood_Analysis_{self.timestamp}.xlsx')
        wb.save(output_path)
        return output_path

    def _create_overview_sheet(self, wb: Workbook, results: dict):
        ws = wb.active
        ws.title = "Overview"

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15

        row = 2

        ws.merge_cells(f'B{row}:D{row}')
        ws[f'B{row}'] = "Urban Flood Comprehensive Analysis Report"
        ws[f'B{row}'].font = self.fonts['title']
        row += 2

        ws[f'B{row}'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws[f'B{row}'].font = self.fonts['small']
        row += 1
        ws[f'B{row}'] = f"Version: 3.1.0"
        ws[f'B{row}'].font = self.fonts['small']
        row += 3

        ws[f'B{row}'] = "Data Summary"
        ws[f'B{row}'].font = self.fonts['section']
        row += 1

        summary = results.get('summary', {})
        summary_items = [
            ("Total Records Processed", f"{summary.get('total_rows', 0):,}"),
            ("Valid Records Analyzed", f"{summary.get('valid_rows', 0):,}"),
            ("Countries/Regions", f"{summary.get('countries_count', 0)}"),
            ("Continents", f"{summary.get('continents_count', 0)}"),
            ("Income Levels", f"{summary.get('income_levels_count', 0)}"),
            ("Built-up Age Categories", f"{summary.get('ages_count', 0)}")
        ]

        for label, value in summary_items:
            ws[f'B{row}'] = f"  - {label}:"
            ws[f'B{row}'].font = self.fonts['body']
            ws[f'C{row}'] = value
            ws[f'C{row}'].font = Font(bold=True, size=10, color='1a73e8')
            row += 1

        row += 2

        ws[f'B{row}'] = "Analysis Contents Directory"
        ws[f'B{row}'].font = self.fonts['section']
        row += 2

        headers = ['Section', 'Sheet Name', 'Description', 'Records']
        for col, header in enumerate(headers, 2):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.fonts['header']
            cell.fill = self.fills['header']
            cell.border = self.border
            cell.alignment = self.align_center
        row += 1

        directory = [
            ("A. Area Statistics", "A_Area_Stats",
             "Built-up area by development status and Built-up Age (includes All)",
             len(results.get('area', {}).get('by_dev_age', []))),

            ("B1. Risk by Development", "B1_Risk_DevStatus",
             "Flood risk index by development status (All/Developed/Developing) with n, σ, SE",
             len(results.get('risk', {}).get('by_dev', []))),

            ("B2. Risk by Income", "B2_Risk_Income",
             "Flood risk index by income classification (includes All) with n, σ, SE",
             len(results.get('risk', {}).get('by_income', []))),

            ("B3. Risk by Continent", "B3_Risk_Continent",
             "Flood risk index by continent (includes All) with n, σ, SE",
             len(results.get('risk', {}).get('by_continent', []))),

            ("B4. Risk by Country", "B4_Risk_Country",
             "Flood risk index by country/region (includes All) with n, σ, SE",
             len(results.get('risk', {}).get('by_country', []))),

            ("C1. Env by Development", "C1_Env_DevStatus",
             "Environmental factors by development status (includes All) with n, σ, SE",
             len(results.get('environment', {}).get('by_dev', []))),

            ("C2. Env by Income", "C2_Env_Income",
             "Environmental factors by income level (includes All) with n, σ, SE",
             len(results.get('environment', {}).get('by_income', []))),

            ("C3. Env by Continent", "C3_Env_Continent",
             "Environmental factors by continent (includes All) with n, σ, SE",
             len(results.get('environment', {}).get('by_continent', []))),

            ("C4. Env by Country", "C4_Env_Country",
             "Environmental factors by country/region (includes All) with n, σ, SE",
             len(results.get('environment', {}).get('by_country', []))),

            ("D1. Global Exposure Ratio", "D1_Global_ExpRatio",
             "Global population exposure ratio by Built-up Age",
             len(results.get('exposure', {}).get('global_ratio', []))),

            ("D2. Exposure by Country-Year", "D2_Exp_Country_Year",
             "Total exposure population by country and year (1975-2025)",
             len(results.get('exposure', {}).get('by_country_year', []))),

            ("D3. Exposure Ratio by Country", "D3_ExpRatio_Country",
             "Exposure ratio by country/region and Built-up Age (includes All)",
             len(results.get('exposure', {}).get('by_country_ratio', []))),

            ("D4. Exposure Ratio by Income", "D4_ExpRatio_Income",
             "Exposure ratio by income level and Built-up Age (includes All)",
             len(results.get('exposure', {}).get('by_income_ratio', []))),

            ("D5. Exposure Ratio by Continent", "D5_ExpRatio_Continent",
             "Exposure ratio by continent and Built-up Age (includes All)",
             len(results.get('exposure', {}).get('by_continent_ratio', []))),

            ("D6. Exposure Ratio by DevStatus", "D6_ExpRatio_DevStatus",
             "Exposure ratio by development status and Built-up Age (includes All)",
             len(results.get('exposure', {}).get('by_dev_ratio', []))),
        ]

        for i, (section, sheet, desc, count) in enumerate(directory):
            for col, val in enumerate([section, sheet, desc, count], 2):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = self.fonts['body']
                cell.border = self.border
                cell.alignment = self.align_left if col < 5 else self.align_center
                if i % 2 == 1:
                    cell.fill = self.fills['alt_row']
            row += 1

        row += 2

        ws[f'B{row}'] = "Notes"
        ws[f'B{row}'].font = self.fonts['section']
        row += 1

        notes = [
            "- Exposure Ratio = Sum(Exposed Population) / Sum(Total Population) x 100%",
            "- Built-up Age: Years since urban development (5, 10, 15, ... 50)",
            "- Risk Index: Composite flood risk indicator (higher = more risk)",
            "- Environmental Factors: Hazard Index, Slope, DEM, Distance from River",
            "- Cell Area: 0.01 km2 per grid cell",
            "- n: Sample count (number of valid observations)",
            "- σ (Std): Standard deviation",
            "- SE (σ/√n): Standard Error = Standard Deviation / Square Root of n",
            "- 'All' category: Aggregated statistics for all data in each grouping"
        ]

        for note in notes:
            ws[f'B{row}'] = note
            ws[f'B{row}'].font = self.fonts['small']
            row += 1

        ws.freeze_panes = 'B1'

    def _create_area_sheet(self, wb: Workbook, area_data: dict):
        ws = wb.create_sheet("A_Area_Stats")

        data = area_data.get('by_dev_age', [])
        if not data:
            ws['A1'] = "No data available"
            return

        headers = ['Category', 'Built-up Age', 'Grid Count', 'Area (km2)', 'Ratio (%)']
        self._write_data_table(ws, headers, data,
                               ['Category', 'Built_Up_Age', 'Grid_Count', 'Area_km2', 'Ratio_Percent'])

    def _create_risk_sheets(self, wb: Workbook, risk_data: dict):
        """Create Risk Index sheets with n and SE (σ/√n) columns"""
        sheet_mapping = {
            'by_dev': 'B1_Risk_DevStatus',
            'by_income': 'B2_Risk_Income',
            'by_continent': 'B3_Risk_Continent',
            'by_country': 'B4_Risk_Country'
        }

        # Updated headers to include n and SE
        headers = ['Category', 'Built-up Age', 'Mean (β)', 'Std (σ)',
                   'Min', 'Max', 'n', 'SE (σ/√n)']
        keys = ['Category', 'Built_Up_Age', 'Risk_Mean', 'Risk_Std',
                'Risk_Min', 'Risk_Max', 'Sample_Count', 'Risk_SE']

        for data_key, sheet_name in sheet_mapping.items():
            data = risk_data.get(data_key, [])
            ws = wb.create_sheet(sheet_name)
            if data:
                self._write_data_table(ws, headers, data, keys)
            else:
                ws['A1'] = "No data available"

    def _create_env_sheets(self, wb: Workbook, env_data: dict):
        """Create Environmental Factor sheets with n and SE for each factor"""
        sheet_mapping = {
            'by_dev': 'C1_Env_DevStatus',
            'by_income': 'C2_Env_Income',
            'by_continent': 'C3_Env_Continent',
            'by_country': 'C4_Env_Country'
        }

        # Updated headers to include n and SE for each environmental factor
        headers = ['Category', 'Built-up Age',
                   'Hazard Mean', 'Hazard Std', 'Hazard n', 'Hazard SE',
                   'Slope Mean', 'Slope Std', 'Slope n', 'Slope SE',
                   'DEM Mean', 'DEM Std', 'DEM n', 'DEM SE',
                   'River Dist Mean', 'River Dist Std', 'River Dist n', 'River Dist SE']

        keys = ['Category', 'Built_Up_Age',
                'Hazard_Index_EAD_Mean', 'Hazard_Index_EAD_Std', 'Hazard_Index_EAD_n', 'Hazard_Index_EAD_SE',
                'Slope_Mean', 'Slope_Std', 'Slope_n', 'Slope_SE',
                'DEM_Mean', 'DEM_Std', 'DEM_n', 'DEM_SE',
                'Distance_from_River_Mean', 'Distance_from_River_Std', 'Distance_from_River_n',
                'Distance_from_River_SE']

        for data_key, sheet_name in sheet_mapping.items():
            data = env_data.get(data_key, [])
            ws = wb.create_sheet(sheet_name)
            if data:
                self._write_data_table(ws, headers, data, keys)
            else:
                ws['A1'] = "No data available"

    def _create_exposure_sheets(self, wb: Workbook, exposure_data: dict):
        # D1. Global Exposure Ratio
        ws = wb.create_sheet("D1_Global_ExpRatio")
        data = exposure_data.get('global_ratio', [])
        if data:
            headers = ['Built-up Age', 'Total Population', 'Total Exposure', 'Exposure Ratio (%)']
            keys = ['Built_Up_Age', 'Total_Population', 'Total_Exposure', 'Exposure_Ratio_Percent']
            self._write_data_table(ws, headers, data, keys)

        # D2. Country Year Exposure
        ws = wb.create_sheet("D2_Exp_Country_Year")
        country_data = exposure_data.get('by_country_year', [])
        if country_data:
            headers = ['Rank', 'Country'] + [str(y) for y in POP_YEARS]
            row_num = 1
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = self.fonts['header']
                cell.fill = self.fills['header']
                cell.border = self.border
                cell.alignment = self.align_center

            for rank, item in enumerate(country_data, 1):
                row_num += 1
                ws.cell(row=row_num, column=1, value=rank).border = self.border
                ws.cell(row=row_num, column=2, value=item['Country']).border = self.border
                for i, year in enumerate(POP_YEARS):
                    val = item['Years'].get(year, 0)
                    cell = ws.cell(row=row_num, column=3 + i, value=round(val, 0))
                    cell.border = self.border
                    cell.number_format = '#,##0'
                if rank % 2 == 0:
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row_num, column=col).fill = self.fills['alt_row']

            self._auto_column_width(ws, len(headers))
            ws.freeze_panes = 'C2'

        # D3. Country Exposure Ratio (includes "All")
        ws = wb.create_sheet("D3_ExpRatio_Country")
        data = exposure_data.get('by_country_ratio', [])
        if data:
            headers = ['Country', 'Built-up Age', 'Total Population',
                       'Total Exposure', 'Exposure Ratio (%)']
            keys = ['Category', 'Built_Up_Age', 'Total_Population',
                    'Total_Exposure', 'Exposure_Ratio_Percent']
            self._write_data_table(ws, headers, data, keys)

        # D4. Income Exposure Ratio (includes "All")
        ws = wb.create_sheet("D4_ExpRatio_Income")
        data = exposure_data.get('by_income_ratio', [])
        if data:
            headers = ['Income Level', 'Built-up Age', 'Total Population',
                       'Total Exposure', 'Exposure Ratio (%)']
            keys = ['Category', 'Built_Up_Age', 'Total_Population',
                    'Total_Exposure', 'Exposure_Ratio_Percent']
            self._write_data_table(ws, headers, data, keys)

        # D5. Continent Exposure Ratio (includes "All")
        ws = wb.create_sheet("D5_ExpRatio_Continent")
        data = exposure_data.get('by_continent_ratio', [])
        if data:
            headers = ['Continent', 'Built-up Age', 'Total Population',
                       'Total Exposure', 'Exposure Ratio (%)']
            keys = ['Category', 'Built_Up_Age', 'Total_Population',
                    'Total_Exposure', 'Exposure_Ratio_Percent']
            self._write_data_table(ws, headers, data, keys)

        # D6. Development Status Exposure Ratio (includes "All")
        ws = wb.create_sheet("D6_ExpRatio_DevStatus")
        data = exposure_data.get('by_dev_ratio', [])
        if data:
            headers = ['Development Status', 'Built-up Age', 'Total Population',
                       'Total Exposure', 'Exposure Ratio (%)']
            keys = ['Category', 'Built_Up_Age', 'Total_Population',
                    'Total_Exposure', 'Exposure_Ratio_Percent']
            self._write_data_table(ws, headers, data, keys)

    def _write_data_table(self, ws, headers: list, data: list, keys: list):
        """Write data table with special formatting for 'All' category rows"""
        if not data:
            ws['A1'] = "No data available"
            return

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.fonts['header']
            cell.fill = self.fills['header']
            cell.border = self.border
            cell.alignment = self.align_center

        for row_idx, item in enumerate(data, 2):
            is_all_row = item.get('Category', '') == 'All'

            for col_idx, key in enumerate(keys, 1):
                value = item.get(key, '')

                if isinstance(value, float):
                    if np.isnan(value):
                        value = 'N/A'
                    elif abs(value) >= 1000000:
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.number_format = '#,##0'
                    elif abs(value) < 0.0001 and value != 0:
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.number_format = '0.00E+00'
                    else:
                        cell = ws.cell(row=row_idx, column=col_idx, value=round(value, 6))
                        cell.number_format = '#,##0.000000'
                else:
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)

                if isinstance(value, str):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)

                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = self.border
                cell.alignment = self.align_center if col_idx > 1 else self.align_left

                # Special formatting for "All" category rows
                if is_all_row:
                    cell.fill = self.fills['all_row']
                    cell.font = self.fonts['all_bold']
                else:
                    cell.font = self.fonts['body']
                    if row_idx % 2 == 0:
                        cell.fill = self.fills['alt_row']

        self._auto_column_width(ws, len(headers))
        ws.freeze_panes = 'A2'

    def _auto_column_width(self, ws, num_cols: int):
        for col in range(1, num_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16


# ============================================================================
# GUI Application
# ============================================================================

class Application:
    """Main Application"""

    def __init__(self, root):
        self.root = root
        self.root.title(f"{APP_NAME} v{APP_VERSION}")
        self.root.geometry("1400x950")
        self.root.minsize(1200, 800)
        self.root.configure(bg=Theme.BG_MAIN)

        self.is_running = False
        self.should_stop = False
        self.processing_thread = None

        self.input_file = tk.StringVar(value=DEFAULT_INPUT)
        self.output_dir = tk.StringVar(value=DEFAULT_OUTPUT)
        self.exclude_zero = tk.BooleanVar(value=True)

        self.start_time = None
        self.timer_running = False

        self._build_ui()
        self._log_system(f"{Icons.ROCKET} {APP_NAME} v{APP_VERSION} initialized")
        self._log_system(f"{Icons.INFO} Ready to start analysis...")

    def _build_ui(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TProgressbar', troughcolor=Theme.PROGRESS_BG,
                        background=Theme.PROGRESS_FG, thickness=8)

        main_frame = tk.Frame(self.root, bg=Theme.BG_MAIN)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        main_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=2)
        main_frame.rowconfigure(0, weight=1)

        left_panel = self._build_left_panel(main_frame)
        left_panel.grid(row=0, column=0, sticky='nsew', padx=(0, 15))

        right_panel = self._build_right_panel(main_frame)
        right_panel.grid(row=0, column=1, sticky='nsew')

    def _build_left_panel(self, parent):
        panel = tk.Frame(parent, bg=Theme.BG_CARD, highlightbackground=Theme.BORDER,
                         highlightthickness=1, relief='flat')
        inner = tk.Frame(panel, bg=Theme.BG_CARD, padx=25, pady=25)
        inner.pack(fill=tk.BOTH, expand=True)

        title_frame = tk.Frame(inner, bg=Theme.BG_CARD)
        title_frame.pack(fill=tk.X, pady=(0, 25))

        tk.Label(title_frame, text=f"{Icons.CHART} Analysis Configuration",
                 font=('Segoe UI', 18, 'bold'),
                 bg=Theme.BG_CARD, fg=Theme.PRIMARY).pack(anchor='w')
        tk.Label(title_frame, text="Urban Flood Comprehensive Analysis v3.1",
                 font=('Segoe UI', 10), bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(anchor='w')

        file_frame = tk.LabelFrame(inner, text=f"{Icons.FOLDER} File Paths",
                                   font=('Segoe UI', 11, 'bold'),
                                   bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY, padx=15, pady=15)
        file_frame.pack(fill=tk.X, pady=(0, 20))

        tk.Label(file_frame, text="Input File (Parquet):",
                 font=('Segoe UI', 10), bg=Theme.BG_CARD).pack(anchor='w', pady=(0, 5))
        input_row = tk.Frame(file_frame, bg=Theme.BG_CARD)
        input_row.pack(fill=tk.X, pady=(0, 15))

        self.input_entry = tk.Entry(input_row, textvariable=self.input_file,
                                    font=('Consolas', 9), width=45)
        self.input_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(input_row, text="Browse", command=self._browse_input,
                  font=('Segoe UI', 9), bg=Theme.PRIMARY_LIGHT,
                  fg=Theme.PRIMARY, relief='flat', padx=15).pack(side=tk.LEFT, padx=(10, 0))

        tk.Label(file_frame, text="Output Directory:",
                 font=('Segoe UI', 10), bg=Theme.BG_CARD).pack(anchor='w', pady=(0, 5))
        output_row = tk.Frame(file_frame, bg=Theme.BG_CARD)
        output_row.pack(fill=tk.X)

        self.output_entry = tk.Entry(output_row, textvariable=self.output_dir,
                                     font=('Consolas', 9), width=45)
        self.output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(output_row, text="Browse", command=self._browse_output,
                  font=('Segoe UI', 9), bg=Theme.PRIMARY_LIGHT,
                  fg=Theme.PRIMARY, relief='flat', padx=15).pack(side=tk.LEFT, padx=(10, 0))

        stats_frame = tk.LabelFrame(inner, text=f"{Icons.STATS} Statistics Contents (v3.1)",
                                    font=('Segoe UI', 11, 'bold'),
                                    bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY, padx=15, pady=15)
        stats_frame.pack(fill=tk.X, pady=(0, 20))

        stats_items = [
            (f"{Icons.BUILDING} Built-up Area Statistics", "By Dev Status x Built-up Age (includes All)"),
            (f"{Icons.WARNING} Flood Risk Index Statistics", "By Dev/Income/Continent/Country (with n, σ, SE)"),
            (f"{Icons.MAP} Environmental Factor Statistics", "Hazard/Slope/DEM/River Dist (with n, σ, SE)"),
            (f"{Icons.PEOPLE} Population Exposure Statistics", "Total Exposure + Ratio by Groups (includes All)"),
        ]

        for title, desc in stats_items:
            item_frame = tk.Frame(stats_frame, bg=Theme.BG_CARD)
            item_frame.pack(fill=tk.X, pady=3)
            tk.Label(item_frame, text=title, font=('Segoe UI', 10),
                     bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY).pack(anchor='w')
            tk.Label(item_frame, text=f"    {desc}", font=('Segoe UI', 9),
                     bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(anchor='w')

        option_frame = tk.LabelFrame(inner, text=f"{Icons.GEAR} Options",
                                     font=('Segoe UI', 11, 'bold'),
                                     bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY, padx=15, pady=15)
        option_frame.pack(fill=tk.X, pady=(0, 20))

        tk.Checkbutton(option_frame, text="Exclude Zero Values (Recommended)", variable=self.exclude_zero,
                       font=('Segoe UI', 10), bg=Theme.BG_CARD,
                       activebackground=Theme.BG_CARD).pack(anchor='w')

        btn_frame = tk.Frame(inner, bg=Theme.BG_CARD)
        btn_frame.pack(fill=tk.X, pady=(10, 0))

        self.start_btn = tk.Button(btn_frame, text=f"{Icons.PLAY} START ANALYSIS",
                                   command=self._start_processing,
                                   font=('Segoe UI', 12, 'bold'),
                                   bg=Theme.SUCCESS, fg='white',
                                   relief='flat', padx=25, pady=12,
                                   cursor='hand2')
        self.start_btn.pack(side=tk.LEFT, padx=(0, 15))

        self.stop_btn = tk.Button(btn_frame, text=f"{Icons.STOP} STOP",
                                  command=self._stop_processing,
                                  font=('Segoe UI', 12, 'bold'),
                                  bg=Theme.ERROR, fg='white',
                                  relief='flat', padx=25, pady=12,
                                  state=tk.DISABLED, cursor='hand2')
        self.stop_btn.pack(side=tk.LEFT)

        return panel

    def _build_right_panel(self, parent):
        panel = tk.Frame(parent, bg=Theme.BG_CARD, highlightbackground=Theme.BORDER,
                         highlightthickness=1, relief='flat')
        inner = tk.Frame(panel, bg=Theme.BG_CARD, padx=20, pady=20)
        inner.pack(fill=tk.BOTH, expand=True)

        progress_frame = tk.Frame(inner, bg=Theme.BG_CARD)
        progress_frame.pack(fill=tk.X, pady=(0, 15))

        status_row = tk.Frame(progress_frame, bg=Theme.BG_CARD)
        status_row.pack(fill=tk.X, pady=(0, 10))

        self.status_label = tk.Label(status_row, text=f"{Icons.INFO} Ready",
                                     font=('Segoe UI', 12, 'bold'),
                                     bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY)
        self.status_label.pack(side=tk.LEFT)

        self.timer_label = tk.Label(status_row, text="00:00:00",
                                    font=('Consolas', 12),
                                    bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED)
        self.timer_label.pack(side=tk.RIGHT)

        self.progress = ttk.Progressbar(progress_frame, mode='determinate', style='TProgressbar')
        self.progress.pack(fill=tk.X, pady=(0, 5))

        self.progress_label = tk.Label(progress_frame, text="0%",
                                       font=('Segoe UI', 9),
                                       bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED)
        self.progress_label.pack()

        stats_outer = tk.Frame(inner, bg=Theme.BG_CARD)
        stats_outer.pack(fill=tk.X, pady=(0, 15))

        self.stat_labels = {}
        stat_items = [
            ('rows', f'{Icons.DATA} Rows:', '0'),
            ('chunks', f'{Icons.LAYER} Chunks:', '0'),
            ('countries', f'{Icons.GLOBE} Countries:', '0'),
            ('sheets', f'{Icons.FILE} Sheets:', '0')
        ]

        for key, label, default in stat_items:
            frame = tk.Frame(stats_outer, bg=Theme.BG_CARD, padx=15, pady=8)
            frame.pack(side=tk.LEFT, expand=True)

            tk.Label(frame, text=label, font=('Segoe UI', 9),
                     bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack()
            self.stat_labels[key] = tk.Label(frame, text=default,
                                             font=('Segoe UI', 14, 'bold'),
                                             bg=Theme.BG_CARD, fg=Theme.PRIMARY)
            self.stat_labels[key].pack()

        log_frame = tk.Frame(inner, bg=Theme.BG_LOG)
        log_frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(log_frame, text=f" {Icons.DATA} Processing Log",
                 font=('Segoe UI', 10, 'bold'),
                 bg=Theme.BG_DARK, fg=Theme.TEXT_LIGHT,
                 anchor='w', padx=10, pady=8).pack(fill=tk.X)

        self.log_text = scrolledtext.ScrolledText(
            log_frame, wrap=tk.WORD, bg=Theme.BG_LOG, fg='#e8e8e8',
            font=('Consolas', 10), insertbackground='white',
            relief='flat', padx=15, pady=10
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)

        self.log_text.tag_configure('INFO', foreground='#e8e8e8')
        self.log_text.tag_configure('SUCCESS', foreground='#4ade80')
        self.log_text.tag_configure('WARNING', foreground='#fbbf24')
        self.log_text.tag_configure('ERROR', foreground='#f87171')
        self.log_text.tag_configure('HEADER', foreground='#60a5fa', font=('Consolas', 11, 'bold'))
        self.log_text.tag_configure('DETAIL', foreground='#9ca3af')
        self.log_text.tag_configure('SYSTEM', foreground='#a78bfa')

        return panel

    def _browse_input(self):
        path = filedialog.askopenfilename(
            title="Select Parquet File",
            filetypes=[("Parquet files", "*.parquet"), ("All files", "*.*")]
        )
        if path:
            self.input_file.set(path)

    def _browse_output(self):
        path = filedialog.askdirectory(title="Select Output Directory")
        if path:
            self.output_dir.set(path)

    def _log(self, message: str, tag: str = 'INFO'):
        self.log_text.insert(tk.END, f"{message}\n", tag)
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def _log_system(self, message: str):
        self._log(message, 'SYSTEM')

    def _log_header(self, message: str):
        self._log(f"\n{'═' * 60}", 'HEADER')
        self._log(f" {message}", 'HEADER')
        self._log(f"{'═' * 60}", 'HEADER')

    def _update_progress(self, value: int, text: str = ""):
        self.progress['value'] = value
        self.progress_label.configure(text=f"{value}% - {text}" if text else f"{value}%")
        self.root.update_idletasks()

    def _update_stats(self, rows=None, chunks=None, countries=None, sheets=None):
        if rows is not None:
            self.stat_labels['rows'].configure(text=f"{rows:,}")
        if chunks is not None:
            self.stat_labels['chunks'].configure(text=str(chunks))
        if countries is not None:
            self.stat_labels['countries'].configure(text=str(countries))
        if sheets is not None:
            self.stat_labels['sheets'].configure(text=str(sheets))
        self.root.update_idletasks()

    def _start_timer(self):
        self.start_time = time.time()
        self.timer_running = True
        self._update_timer()

    def _update_timer(self):
        if self.timer_running:
            elapsed = int(time.time() - self.start_time)
            hours, remainder = divmod(elapsed, 3600)
            minutes, seconds = divmod(remainder, 60)
            self.timer_label.configure(text=f"{hours:02d}:{minutes:02d}:{seconds:02d}")
            self.root.after(1000, self._update_timer)

    def _stop_timer(self):
        self.timer_running = False

    def _start_processing(self):
        input_path = self.input_file.get()
        output_path = self.output_dir.get()

        if not input_path or not os.path.exists(input_path):
            messagebox.showerror("Error", "Please select a valid input file")
            return
        if not output_path or not os.path.isdir(output_path):
            messagebox.showerror("Error", "Please select a valid output directory")
            return

        self.is_running = True
        self.should_stop = False
        self.start_btn.configure(state=tk.DISABLED)
        self.stop_btn.configure(state=tk.NORMAL)

        self.log_text.delete(1.0, tk.END)
        self._update_progress(0, "Initializing...")
        self._update_stats(rows=0, chunks=0, countries=0, sheets=0)

        self.status_label.configure(text=f"{Icons.GEAR} Processing...", fg=Theme.INFO)
        self._start_timer()

        self.processing_thread = threading.Thread(
            target=self._run_analysis,
            args=(input_path, output_path)
        )
        self.processing_thread.start()

    def _stop_processing(self):
        self.should_stop = True
        self.status_label.configure(text=f"{Icons.WARNING} Stopping...", fg=Theme.WARNING)
        self._log(f"\n{Icons.WARNING} Stop requested, please wait...", 'WARNING')

    def _run_analysis(self, input_path: str, output_path: str):
        try:
            self._log_header(f"{Icons.ROCKET} ANALYSIS STARTED")
            self._log(f"{Icons.FILE} Input: {input_path}", 'INFO')
            self._log(f"{Icons.FOLDER} Output: {output_path}", 'INFO')

            self._log(f"\n{Icons.INFO} Opening Parquet file...", 'INFO')
            parquet_file = pq.ParquetFile(input_path)

            schema = parquet_file.schema_arrow
            all_columns = [field.name for field in schema]
            num_row_groups = parquet_file.metadata.num_row_groups
            total_rows = parquet_file.metadata.num_rows

            self._log(f"  {Icons.BULLET} Total columns: {len(all_columns)}", 'DETAIL')
            self._log(f"  {Icons.BULLET} Total rows: {total_rows:,}", 'DETAIL')
            self._log(f"  {Icons.BULLET} Row groups: {num_row_groups}", 'DETAIL')

            if self.should_stop:
                self._complete(False, "User stopped")
                return

            self._log(f"\n{Icons.GEAR} Analyzing required columns...", 'INFO')
            cols_to_read = ['Built_Up_Age']

            optional_cols = {
                'Developed': 'Development Status',
                'Country': 'Country Code',
                'Income_Classification': 'Income Classification',
                'Sovereig': 'Continent Code',
                'Risk_Index': 'Risk Index',
                'Hazard_Index_EAD': 'Hazard Index',
                'Slope': 'Slope',
                'DEM': 'Elevation',
                'Distance_from_River': 'River Distance',
                'POP_2025': 'Population 2025',
                'POP_Exposure_2025': 'Exposure 2025'
            }

            for col, desc in optional_cols.items():
                if col in all_columns:
                    cols_to_read.append(col)
                    self._log(f"  {Icons.CHECK} {col} ({desc})", 'DETAIL')
                else:
                    self._log(f"  {Icons.WARNING} {col} ({desc}) - Missing", 'WARNING')

            pop_cols_found = 0
            for year in POP_YEARS:
                col = f'POP_Exposure_{year}'
                if col in all_columns and col not in cols_to_read:
                    cols_to_read.append(col)
                    pop_cols_found += 1

            if pop_cols_found > 0:
                self._log(f"  {Icons.CHECK} Yearly exposure columns: {pop_cols_found}", 'DETAIL')

            cols_to_read = list(set(cols_to_read))
            self._log(f"\n{Icons.INFO} Total columns to read: {len(cols_to_read)}", 'INFO')

            self._log(f"\n{Icons.CALC} Initializing analyzer (v3.1 with All category)...", 'INFO')
            analyzer = UnifiedDataAnalyzer(
                log_callback=lambda msg: self._log(msg, 'INFO'),
                progress_callback=self._update_progress,
                detail_callback=lambda msg: self._log(msg, 'DETAIL')
            )

            self._update_progress(5, "Starting chunk processing...")

            processed_rows = 0
            valid_rows = 0

            self._log_header(f"{Icons.DATA} CHUNK PROCESSING")

            for idx in range(num_row_groups):
                if self.should_stop:
                    self._complete(False, "User stopped")
                    return

                chunk_start_time = time.time()

                self._log(f"\n{Icons.LAYER} Processing chunk [{idx + 1}/{num_row_groups}]", 'INFO')

                self._log(f"  {Icons.ARROW} Reading data...", 'DETAIL')
                table = parquet_file.read_row_group(idx, columns=cols_to_read)
                df = table.to_pandas()
                chunk_rows = len(df)

                self._log(f"  {Icons.BULLET} Rows: {chunk_rows:,}", 'DETAIL')

                self._log(f"  {Icons.ARROW} Computing statistics...", 'DETAIL')
                valid = analyzer.process_chunk(df, self.exclude_zero.get())

                processed_rows += chunk_rows
                valid_rows += valid

                chunk_time = time.time() - chunk_start_time

                self._log(f"  {Icons.BULLET} Valid rows: {valid:,}", 'DETAIL')
                self._log(f"  {Icons.BULLET} Countries found: {len(analyzer.stats_counters['countries_found'])}",
                          'DETAIL')
                self._log(f"  {Icons.CLOCK} Time: {chunk_time:.2f}s", 'DETAIL')

                progress = 5 + int((processed_rows / total_rows) * 70)
                self._update_progress(progress, f"Processing chunk {idx + 1}/{num_row_groups}")
                self._update_stats(rows=processed_rows, chunks=idx + 1,
                                   countries=len(analyzer.stats_counters['countries_found']))

                del df, table
                gc.collect()

            self._log(f"\n{Icons.SUCCESS} Data processing completed!", 'SUCCESS')
            self._log(f"  {Icons.BULLET} Total rows processed: {processed_rows:,}", 'DETAIL')
            self._log(f"  {Icons.BULLET} Valid data rows: {valid_rows:,}", 'DETAIL')

            if self.should_stop:
                self._complete(False, "User stopped")
                return

            self._log_header(f"{Icons.CALC} COMPUTING FINAL STATISTICS")
            self._update_progress(80, "Computing final statistics...")

            self._log(f"{Icons.ARROW} Summarizing area statistics (with All)...", 'INFO')
            self._log(f"{Icons.ARROW} Computing risk index statistics (with n, σ, SE)...", 'INFO')
            self._log(f"{Icons.ARROW} Computing environmental factor statistics (with n, σ, SE)...", 'INFO')
            self._log(f"{Icons.ARROW} Computing population exposure ratios (with All)...", 'INFO')

            results = analyzer.calculate_final_results()

            self._log(f"\n{Icons.SUCCESS} Statistics computation completed!", 'SUCCESS')

            self._log_header(f"{Icons.SAVE} GENERATING EXCEL REPORT")
            self._update_progress(90, "Generating Excel report...")

            self._log(f"{Icons.ARROW} Creating workbook...", 'INFO')
            report_gen = ExcelReportGenerator(output_path)

            self._log(f"{Icons.ARROW} Writing Overview sheet...", 'DETAIL')
            self._log(f"{Icons.ARROW} Writing Area Statistics sheet...", 'DETAIL')
            self._log(f"{Icons.ARROW} Writing Risk Index sheets (4) with n and SE...", 'DETAIL')
            self._log(f"{Icons.ARROW} Writing Environmental Factor sheets (4) with n and SE...", 'DETAIL')
            self._log(f"{Icons.ARROW} Writing Population Exposure sheets (6)...", 'DETAIL')

            excel_path = report_gen.generate_report(results)

            self._update_stats(sheets=16)

            self._log(f"\n{Icons.SUCCESS} Report generated successfully!", 'SUCCESS')
            self._log(f"  {Icons.FILE} {excel_path}", 'INFO')

            self._update_progress(100, "Complete!")
            self._complete(True, f"Analysis completed!\n\nReport saved to:\n{excel_path}")

        except Exception as e:
            self._log(f"\n{Icons.ERROR} Error: {str(e)}", 'ERROR')
            traceback.print_exc()
            self._complete(False, str(e))

    def _complete(self, success: bool, message: str):
        self._stop_timer()
        self.is_running = False
        self.start_btn.configure(state=tk.NORMAL)
        self.stop_btn.configure(state=tk.DISABLED)

        if success:
            self.status_label.configure(text=f"{Icons.SUCCESS} Complete", fg=Theme.SUCCESS)
            self._log_header(f"{Icons.SUCCESS} ANALYSIS COMPLETED")
            messagebox.showinfo("Success", message)
        else:
            self.status_label.configure(text=f"{Icons.ERROR} Failed", fg=Theme.ERROR)
            if "stopped" not in message.lower():
                messagebox.showerror("Error", message)


def main():
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except:
        pass

    root = tk.Tk()
    app = Application(root)

    def on_closing():
        if app.is_running:
            if messagebox.askokcancel("Confirm", "Processing is running. Are you sure you want to quit?"):
                app.should_stop = True
                root.after(500, root.destroy)
        else:
            root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_closing)
    root.mainloop()


if __name__ == "__main__":
    main()